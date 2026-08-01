from __future__ import annotations
import io, math, os, json
from typing import List, Tuple, Dict, Optional
from contextlib import asynccontextmanager

import asyncio
import httpx
import aiosqlite
from fastapi import FastAPI, Query, HTTPException, Request
from fastapi.exceptions import RequestValidationError
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, Response as FastAPIResponse, JSONResponse
from fastapi.encoders import jsonable_encoder
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware
from pydantic import BaseModel, Field, field_validator
from xml.sax.saxutils import escape as xml_escape
import pandas as pd

LOCATIONIQ_KEY  = os.getenv("LOCATIONIQ_KEY", "") or os.getenv("LOCATIONIQ_TOKEN", "")
CONTACT_EMAIL   = os.getenv("CONTACT_EMAIL", "")
USER_AGENT      = f"TripLogApp/9.0 (contact:{CONTACT_EMAIL or 'none'})"
DB_PATH         = os.getenv("CACHE_DB", "cache.db")
ALLOWED_ORIGINS = [o.strip() for o in os.getenv("ALLOWED_ORIGINS", "*").split(",")]
NO_CACHE_HEADERS = {"Cache-Control": "no-cache, no-store, must-revalidate"}
MAX_ROUTE_POINTS = 50
MAX_EXPORT_ROWS = 10000
EXCEL_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _json_nocache(data) -> JSONResponse:
    return JSONResponse(content=data, headers=NO_CACHE_HEADERS)

limiter = Limiter(key_func=get_remote_address, default_limits=["60/minute"])


@asynccontextmanager
async def lifespan(app: FastAPI):
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("CREATE TABLE IF NOT EXISTS geo_cache   (key TEXT PRIMARY KEY, value TEXT NOT NULL)")
        await db.execute("CREATE TABLE IF NOT EXISTS route_cache (key TEXT PRIMARY KEY, value TEXT NOT NULL)")
        await db.commit()
    yield


def _finite_safe(o):
    """Înlocuiește NaN/Infinity din payloaduri (inclusiv din contextul erorilor de validare)
    cu stringuri, astfel încât răspunsurile de eroare să poată fi serializate ca JSON."""
    if isinstance(o, float) and not math.isfinite(o):
        return "NaN" if math.isnan(o) else ("Infinity" if o > 0 else "-Infinity")
    if isinstance(o, dict):
        return {k: _finite_safe(v) for k, v in o.items()}
    if isinstance(o, (list, tuple)):
        return [_finite_safe(v) for v in o]
    return o


app = FastAPI(title="TripLog", lifespan=lifespan)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)


@app.exception_handler(RequestValidationError)
async def _validation_exception_handler(request: Request, exc: RequestValidationError):
    return JSONResponse(
        status_code=422,
        content={"detail": jsonable_encoder(_finite_safe(exc.errors()))},
        headers=NO_CACHE_HEADERS,
    )


app.add_middleware(SlowAPIMiddleware)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
)


def km_round(x: float) -> float:
    return math.floor(x * 10 + 0.5) / 10


# ── Geocodare ────────────────────────────────────────────────────────────────

async def _try_locationiq(client: httpx.AsyncClient, q: str, limit: int) -> Optional[list]:
    if not LOCATIONIQ_KEY:
        return None
    try:
        r = await client.get(
            "https://us1.locationiq.com/v1/search",
            params={"key": LOCATIONIQ_KEY, "q": q, "format": "json",
                    "normalizecity": 1, "limit": str(limit), "accept-language": "ro"},
        )
        r.raise_for_status()
        js = r.json()
        return ([{"lat": float(it["lat"]), "lon": float(it["lon"]),
                  "display": it.get("display_name", q)} for it in js]
                if isinstance(js, list) else None)
    except Exception:
        return None


async def _try_nominatim(client: httpx.AsyncClient, q: str, limit: int) -> Optional[list]:
    try:
        r = await client.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": q, "format": "json", "limit": limit, "accept-language": "ro"},
        )
        r.raise_for_status()
        js = r.json()
        return ([{"lat": float(it["lat"]), "lon": float(it["lon"]),
                  "display": it.get("display_name", q)} for it in js]
                if isinstance(js, list) else None)
    except Exception:
        return None


async def _try_mapsco(client: httpx.AsyncClient, q: str, limit: int) -> Optional[list]:
    try:
        r = await client.get("https://geocode.maps.co/search",
                             params={"q": q, "limit": str(limit)})
        r.raise_for_status()
        js = r.json()
        if not isinstance(js, list):
            return None
        return ([{"lat": float(it["lat"]), "lon": float(it["lon"]),
                  "display": it.get("display_name") or it.get("name") or q}
                 for it in js if it.get("lat") is not None and it.get("lon") is not None] or None)
    except Exception:
        return None


@app.get("/api/geocode")
async def geocode(q: str = Query(..., min_length=3),
                  limit: int = Query(6, ge=1, le=10)):
    key = f"{q}|{limit}"
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute("SELECT value FROM geo_cache WHERE key=?", (key,)) as cur:
            row = await cur.fetchone()
            if row:
                return _json_nocache({"results": json.loads(row[0])})

    async with httpx.AsyncClient(headers={"User-Agent": USER_AGENT}, timeout=10.0) as client:
        result = (await _try_locationiq(client, q, limit)
                  or await _try_nominatim(client, q, limit)
                  or await _try_mapsco(client, q, limit)
                  or [])

    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("INSERT OR REPLACE INTO geo_cache VALUES (?,?)", (key, json.dumps(result)))
        await db.commit()
    return _json_nocache({"results": result})


@app.get("/api/reverse")
async def reverse_geocode(lat: float = Query(...),
                          lon: float = Query(...)):
    if not (math.isfinite(lat) and -90 <= lat <= 90):
        raise HTTPException(400, "Latitudine invalidă.")
    if not (math.isfinite(lon) and -180 <= lon <= 180):
        raise HTTPException(400, "Longitudine invalidă.")
    key  = f"rev|{lat:.5f},{lon:.5f}"
    # Cache lookup
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute("SELECT value FROM geo_cache WHERE key=?", (key,)) as cur:
            row = await cur.fetchone()
            if row:
                return _json_nocache(json.loads(row[0]))

    display = None
    async with httpx.AsyncClient(headers={"User-Agent": USER_AGENT}, timeout=10.0) as client:
        # 1) LocationIQ (dacă există cheia)
        if LOCATIONIQ_KEY:
            try:
                r = await client.get(
                    "https://us1.locationiq.com/v1/reverse",
                    params={"key": LOCATIONIQ_KEY, "lat": lat, "lon": lon,
                            "format": "json", "accept-language": "ro"},
                )
                if r.status_code == 200:
                    d = r.json()
                    display = d.get("display_name")
            except Exception:
                pass

        # 2) Nominatim fallback
        if not display:
            try:
                r = await client.get(
                    "https://nominatim.openstreetmap.org/reverse",
                    params={"format": "json", "lat": lat, "lon": lon, "accept-language": "ro"},
                )
                if r.status_code == 200:
                    d = r.json()
                    display = d.get("display_name")
            except Exception:
                pass

    result = {"display": display or f"{lat:.5f}, {lon:.5f}", "lat": lat, "lon": lon}
    # Salvează în cache doar dacă avem o adresă reală
    if display:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute("INSERT OR REPLACE INTO geo_cache VALUES (?,?)",
                             (key, json.dumps(result)))
            await db.commit()
    return _json_nocache(result)


# ── Rutare ───────────────────────────────────────────────────────────────────

class RouteRequest(BaseModel):
    points: List[Tuple[float, float]] = Field(default_factory=list)

    @field_validator("points")
    @classmethod
    def _check_points(cls, v):
        if not (2 <= len(v) <= MAX_ROUTE_POINTS):
            raise ValueError(f"Numărul de puncte trebuie să fie între 2 și {MAX_ROUTE_POINTS}.")
        for lat, lon in v:
            if not math.isfinite(lat) or not math.isfinite(lon):
                raise ValueError("Coordonate invalide.")
            if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
                raise ValueError("Coordonate în afara intervalului.")
        return v


@app.post("/api/route")
async def route(body: RouteRequest):
    key = "v2|" + "|".join(f"{la:.5f},{lo:.5f}" for la, lo in body.points)

    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute("SELECT value FROM route_cache WHERE key=?", (key,)) as cur:
            row = await cur.fetchone()
            if row:
                return _json_nocache(json.loads(row[0]))

    coord_str = ";".join(f"{lo},{la}" for la, lo in body.points)
    url = (f"https://router.project-osrm.org/route/v1/driving/{coord_str}"
           f"?overview=full&alternatives=false&steps=false&geometries=geojson")
    try:
        async with httpx.AsyncClient(headers={"User-Agent": USER_AGENT}, timeout=20.0) as client:
            r = await client.get(url)
        if r.status_code == 400:
            raise HTTPException(400, r.json().get("message", "Nu s-a găsit rută."))
        r.raise_for_status()
        data = r.json()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(503, f"Eroare rețea OSRM: {e}")

    routes = data.get("routes") or []
    if not routes:
        raise HTTPException(404, "Nicio rută găsită.")

    legs     = routes[0].get("legs", [])
    legs_km  = [km_round(leg["distance"] / 1000) for leg in legs]
    legs_min = [round(leg.get("duration", 0) / 60, 1) for leg in legs]
    geom     = routes[0].get("geometry", {})
    raw      = geom.get("coordinates", []) if geom.get("type") == "LineString" else []
    coords   = raw[::3]
    if raw and (not coords or coords[-1] != raw[-1]):
        coords.append(raw[-1])

    result = {"legs_km": legs_km, "legs_min": legs_min, "coords": coords}
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("INSERT OR REPLACE INTO route_cache VALUES (?,?)", (key, json.dumps(result)))
        await db.commit()
    return _json_nocache(result)


# ── Export Excel ─────────────────────────────────────────────────────────────

class ExportRequest(BaseModel):
    rows: List[Dict]
    total: float
    date_str: str
    total_col_label: str = "Total KM"


def _sanitize_formula_injection(value):
    """Prefixează cu apostrof valorile text care ar putea fi interpretate ca formule Excel."""
    if isinstance(value, str) and value.startswith(EXCEL_FORMULA_PREFIXES):
        return "'" + value
    return value


def _build_excel(rows, total, total_col_label) -> bytes:
    df   = pd.DataFrame(rows).map(_sanitize_formula_injection)
    bio  = io.BytesIO()
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    with pd.ExcelWriter(bio, engine="openpyxl") as w:
        df.to_excel(w, index=False, header=False, sheet_name="Foaie de parcurs", startrow=1)
        ws   = w.sheets["Foaie de parcurs"]
        fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        hf   = Font(bold=True, color="FFFFFF", size=11)
        ca   = Alignment(horizontal="center", vertical="center")
        for i, col in enumerate(df.columns, 1):
            c = ws.cell(1, i, col); c.font = hf; c.fill = fill; c.alignment = ca
        tc = df.shape[1] + 1
        th = ws.cell(1, tc, total_col_label); th.font = hf; th.fill = fill; th.alignment = ca
        tv = ws.cell(2, tc, float(total))
        tv.font = Font(bold=True, size=12, color="4338CA")
        tv.number_format = "0.0"; tv.alignment = ca
        for ci in range(1, tc + 1):
            mx = max((len(str(ws.cell(r, ci).value or ''))
                      for r in range(1, ws.max_row + 1)), default=10)
            ws.column_dimensions[get_column_letter(ci)].width = min(mx + 4, 55)
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28
    return bio.getvalue()


@app.post("/api/export/excel")
async def export_excel(body: ExportRequest):
    if not body.rows:
        raise HTTPException(400, "Nicio înregistrare de exportat.")
    if len(body.rows) > MAX_EXPORT_ROWS:
        raise HTTPException(400, f"Prea multe rânduri (max {MAX_EXPORT_ROWS}).")
    try:
        data = await asyncio.to_thread(_build_excel, body.rows, body.total, body.total_col_label)
    except Exception as e:
        raise HTTPException(500, str(e))
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=foaie_parcurs_{body.date_str}.xlsx",
                 **NO_CACHE_HEADERS},
    )


# ── Export PDF ────────────────────────────────────────────────────────────────

class PDFRequest(BaseModel):
    rows: List[Dict]
    total: float
    date_str: str
    title: str = "Foaie de parcurs"
    vehicle: str = ""
    driver: str = ""
    total_col_label: str = "Total KM"


def _register_unicode_font() -> str:
    """Înregistrează DejaVu Sans (regular + bold) cu suport Unicode pentru diacritice.
    Returnează 'DejaVuSans' dacă reușește, 'Helvetica' ca fallback."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = "DejaVuSans"
    if font_name in pdfmetrics._fonts:
        return font_name

    base_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(base_dir, "fonts"),                        # bundled în proiect
        "/usr/share/fonts/truetype/dejavu",                     # Debian/Ubuntu
        "/usr/share/fonts/dejavu",                              # Fedora/CentOS
        "/usr/share/fonts/TTF",                                 # Arch
    ]
    for d in candidates:
        reg  = os.path.join(d, "DejaVuSans.ttf")
        bold = os.path.join(d, "DejaVuSans-Bold.ttf")
        if os.path.isfile(reg):
            pdfmetrics.registerFont(TTFont("DejaVuSans",      reg))
            pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", bold if os.path.isfile(bold) else reg))
            pdfmetrics.registerFontFamily(
                "DejaVuSans",
                normal="DejaVuSans", bold="DejaVuSans-Bold",
                italic="DejaVuSans", boldItalic="DejaVuSans-Bold",
            )
            return font_name
    return "Helvetica"   # fallback dacă fontul lipsește


def _build_pdf(body: PDFRequest) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER

    fn      = _register_unicode_font()          # ex. 'DejaVuSans'
    fn_bold = fn + "-Bold" if fn != "Helvetica" else "Helvetica-Bold"

    bio = io.BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles  = getSampleStyleSheet()
    t_style = ParagraphStyle('T', parent=styles['Title'], fontSize=16, spaceAfter=4,
                             fontName=fn_bold, alignment=TA_CENTER,
                             textColor=colors.HexColor('#4338CA'))
    s_style = ParagraphStyle('S', parent=styles['Normal'], fontSize=9, spaceAfter=10,
                             fontName=fn, alignment=TA_CENTER,
                             textColor=colors.HexColor('#64748B'))

    # Stiluri cu word-wrap pentru celule — folosesc fontul Unicode
    hdr_ps   = ParagraphStyle('H',  fontSize=8,  fontName=fn_bold,
                              textColor=colors.white, alignment=TA_CENTER,
                              leading=10, wordWrap='LTR')
    cell_ps  = ParagraphStyle('C',  fontSize=7,  fontName=fn, alignment=TA_CENTER,
                              leading=9,  wordWrap='LTR',
                              textColor=colors.HexColor('#0F172A'))
    total_ps = ParagraphStyle('TC', fontSize=9,  fontName=fn_bold,
                              alignment=TA_CENTER, leading=11,
                              textColor=colors.HexColor('#4338CA'))

    story = [Paragraph(xml_escape(body.title), t_style)]
    meta  = " | ".join(filter(None, [
        body.vehicle and f"Vehicul: {xml_escape(body.vehicle)}",
        body.driver  and f"Șofer: {xml_escape(body.driver)}",
    ]))
    if meta:
        story.append(Paragraph(meta, s_style))
    story.append(Spacer(1, 0.3*cm))

    if body.rows:
        cols = list(body.rows[0].keys())

        # Header + rânduri cu Paragraph → wrap automat
        data = [[Paragraph(xml_escape(c), hdr_ps) for c in cols]
                + [Paragraph(xml_escape(body.total_col_label), hdr_ps)]]
        for i, row in enumerate(body.rows):
            cells = [Paragraph(xml_escape(str(row.get(c, ''))), cell_ps) for c in cols]
            cells.append(Paragraph(f"{body.total:.1f}" if i == 0 else '', total_ps if i == 0 else cell_ps))
            data.append(cells)

        page_w = landscape(A4)[0] - 3*cm

        # Coloane late (adrese, scop) — RO și EN
        wide = {'plecare', 'destinație', 'destinatie', 'departure', 'from',
                'destination', 'to', 'scop', 'purpose'}
        raw_w = [0.22 if any(w in c.lower() for w in wide) else 0.09 for c in cols]
        raw_w.append(0.10)                        # coloana Total
        scale = 1.0 / sum(raw_w)                  # normalizare → suma = page_w
        col_w = [page_w * w * scale for w in raw_w]

        tbl = Table(data, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1,  0), colors.HexColor('#6366F1')),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#E2E8F0')),
            ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING',   (0, 0), (-1, -1), 4),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ]))
        story.append(tbl)

    doc.build(story)
    return bio.getvalue()


@app.post("/api/export/pdf")
async def export_pdf(body: PDFRequest):
    if not body.rows:
        raise HTTPException(400, "Nicio înregistrare de exportat.")
    if len(body.rows) > MAX_EXPORT_ROWS:
        raise HTTPException(400, f"Prea multe rânduri (max {MAX_EXPORT_ROWS}).")
    try:
        data = await asyncio.to_thread(_build_pdf, body)
    except ImportError:
        raise HTTPException(500, "reportlab nu este instalat. Rulați: pip install reportlab")
    except Exception as e:
        raise HTTPException(500, str(e))
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=foaie_parcurs_{body.date_str}.pdf",
                 **NO_CACHE_HEADERS},
    )


# ── Health / SW / Static ──────────────────────────────────────────────────────

@app.get("/health")
@app.head("/health")
async def health():
    return JSONResponse({"status": "ok"}, headers=NO_CACHE_HEADERS)


@app.get("/static/sw.js")
async def service_worker():
    with open("static/sw.js") as f:
        return FastAPIResponse(f.read(), media_type="application/javascript",
                               headers={"Service-Worker-Allowed": "/",
                                        "Cache-Control": "no-cache, no-store, must-revalidate"})


INDEX_PATH = os.path.join("static", "index.html")
if os.path.isdir("static"):
    app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/")
async def root():
    return FileResponse(INDEX_PATH, headers=NO_CACHE_HEADERS)


@app.get("/{full_path:path}")
async def spa(full_path: str):
    if full_path.startswith("api/"):
        raise HTTPException(404, "Not found")
    return FileResponse(INDEX_PATH, headers=NO_CACHE_HEADERS)
