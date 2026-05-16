# -*- coding: utf-8 -*-
"""
Foaie de parcurs - calcul automat km (OSRM gratuit)
v7.0 - Modificări față de v6.5:
  - Cache exclusiv prin @st.cache_data (eliminat cache pe disc, filesystem efemer pe Cloud)
  - Eliminat time.sleep() din thread-ul principal (nu mai blochează UI-ul la requesturi)
  - Favorite per-sesiune cu export/import JSON (nu mai sunt scrise pe serverul Cloud)
  - Adăugate pydeck, gspread, google-auth în requirements.txt
  - Comentarii km_round și PATH_SUBSAMPLE_STEP
"""

from __future__ import annotations
import io, sys, json, time, math
from datetime import date, datetime
from typing import List, Dict, Optional, Tuple

import requests
import pandas as pd

try:
    import streamlit as st
except Exception:
    st = None

# ---------------- Page + CSS ----------------
if st is not None:
    try:
        st.set_page_config(page_title="Foaie de parcurs - calcul automat km", page_icon="🚗", layout="wide")
    except Exception:
        pass
    st.markdown(
        """
        <style>
        #MainMenu, header, footer {visibility:hidden;}
        .block-container {padding-top:.75rem; padding-bottom:5rem; max-width:980px;}

        .card {
          padding:.2rem 0 1rem 0;
          border: none !important;
          border-radius: 0 !important;
          background: transparent !important;
          box-shadow: none !important;
          margin: 0 0 .6rem 0;
        }
        .card :is([data-testid="stMarkdownContainer"], [data-testid="stTextInput"], [data-testid="stSelectbox"], [data-testid="stDateInput"]) {
          margin-top: .25rem !important;
        }
        .card-title {font-weight:700; margin:0; display:inline-block; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;}
        .muted {color:#666; font-size:.9rem}

        h1, h2, h3, h4, h5, h6 { border-bottom: none !important; box-shadow: none !important; padding-bottom: 0 !important; margin-bottom: .4rem !important; }
        [data-testid="stHeading"] hr, [data-testid="stHeading"] div[role="separator"], [data-testid="stHeadingWithDivider"], [data-testid="stMarkdownContainer"] hr, .stHeadingContainer hr { display: none !important; }
        .section { border: none !important; box-shadow: none !important; padding: 0 !important; margin: 0 0 .4rem 0 !important; font-weight: 700; }
        hr { display: none !important; }
        [data-testid="stHeader"] div:after, .block-container h1:after, .block-container h2:after, .block-container h3:after, .block-container h4:after, .block-container h5:after, .block-container h6:after { content: none !important; border: none !important; }

        input, textarea, .stSelectbox div[role="button"], .stSelectbox input {min-height:44px;}
        .stTextInput>div>div, .stSelectbox>div>div { border-radius:10px!important; border:1px solid var(--border,#e6e6e6)!important; }

        .op-row-marker + div [data-testid="stHorizontalBlock"]{ display: grid !important; grid-template-columns: 1fr auto !important; align-items: center !important; gap: .2rem !important; }
        .op-row-marker + div [data-testid="column"]{ width: auto !important; min-width: 0 !important; padding: 0 !important; }
        .op-row-marker + div [data-testid="column"]:first-child{ justify-content: flex-start !important; align-items: center !important; }
        .op-row-marker + div [data-testid="column"]:last-child{ justify-content: flex-end !important; align-items: center !important; }
        .op-row-marker + div .stButton>button{ border-radius:999px!important; width:24px!important; height:24px!important; min-height:24px!important; padding:0!important; line-height:1!important; font-size:14px!important; margin:0!important; }

        @media (max-width: 480px){
          .block-container{padding-left:.5rem; padding-right:.5rem;}
          .op-row-marker + div [data-testid="stHorizontalBlock"]{ gap:.12rem!important; }
          .op-row-marker + div .stButton>button{ width:22px!important; height:22px!important; min-height:22px!important; font-size:13px!important; }
        }

        @media (prefers-color-scheme: dark){
          :root{--bg:#0e1117;--fg:#e6e6e6;--card:#161a23;--muted:#a3a3a3;--border:#2b3040;}
          body{color:var(--fg); background:var(--bg);}
          .card{background:var(--card); border-color:var(--border); box-shadow:none;}
          .stTextInput input,.stSelectbox div[role="button"],.stSelectbox input{ background:var(--card)!important;color:var(--fg)!important;border-radius:10px; }
          ul[role="listbox"]{ background:var(--card)!important;color:var(--fg)!important;border:1px solid var(--border)!important; }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

# ---------------- Constante & Secrets ----------------
APP_TITLE      = "Foaie de parcurs - calcul automat km"
NOMINATIM_URL  = "https://nominatim.openstreetmap.org/search"
MAPSCO_URL     = "https://geocode.maps.co/search"
LOCATIONIQ_URL = "https://us1.locationiq.com/v1/search"

def _secret(name: str, default: str = "") -> str:
    if st is None: return default
    try: return (st.secrets.get(name, default) or default).strip()
    except Exception: return default

CONTACT_EMAIL                = _secret("CONTACT_EMAIL", "")
LOCATIONIQ_KEY               = _secret("LOCATIONIQ_TOKEN", "") or _secret("LOCATIONIQ_KEY", "")
GSPREAD_SERVICE_ACCOUNT_JSON = _secret("GSPREAD_SERVICE_ACCOUNT_JSON", "")
GOOGLE_SHEET_ID              = _secret("GOOGLE_SHEET_ID", "")

USER_AGENT = f"FoaieParcursApp/7.0 ({'mailto:'+CONTACT_EMAIL if CONTACT_EMAIL else 'no-contact'})"

# Câte puncte consecutive din geometria OSRM păstrăm pe hartă.
# 3 = 1 din 3, reduce traficul de rendering fără a distorsiona vizibil traseul.
PATH_SUBSAMPLE_STEP = 3

# ---------------- Utilitare ----------------
def _round5(x: float) -> float:
    return float(f"{x:.5f}")

def km_round(x: float, decimals: int = 1) -> float:
    """
    Rotunjire aritmetică standard (0.5 → sus).
    Python built-in round() folosește banker's rounding (0.5 → par),
    care poate surprinde utilizatorii la valori ca 12.35 km.
    """
    pow10 = 10 ** decimals
    return math.floor(x * pow10 + 0.5) / pow10

def _simplify_coords(coords: List[List[float]], step: int = PATH_SUBSAMPLE_STEP) -> List[List[float]]:
    """Subsamplează coordonatele rutei pentru rendering mai rapid pe hartă."""
    if not coords or step <= 1:
        return coords
    out = coords[::max(1, int(step))]
    if out and coords[-1] != out[-1]:
        out.append(coords[-1])
    return out

# ---------------- Geocodare ----------------
# Cache-ul este gestionat exclusiv prin @st.cache_data (ttl=24h).
# Nu există sleep() – @st.cache_data previne requesturile duplicate pentru aceeași interogare,
# deci nu putem atinge rate limit-ul providerilor în mod normal.

if st is not None:
    @st.cache_data(ttl=24*3600, show_spinner=False)
    def _locationiq_cached(q: str, limit: int, key: str) -> List[Dict]:
        r = requests.get(
            LOCATIONIQ_URL,
            params={"key": key, "q": q, "format": "json", "normalizecity": 1,
                    "limit": str(limit), "accept-language": "ro"},
            headers={"User-Agent": USER_AGENT},
            timeout=12,
        )
        r.raise_for_status()
        js = r.json() if isinstance(r.json(), list) else []
        return [{"lat": float(it["lat"]), "lon": float(it["lon"]),
                 "display": it.get("display_name", q)} for it in js]

    @st.cache_data(ttl=24*3600, show_spinner=False)
    def _nominatim_cached(q: str, limit: int) -> List[Dict]:
        r = requests.get(
            NOMINATIM_URL,
            params={"q": q, "format": "json", "limit": limit, "accept-language": "ro"},
            headers={"User-Agent": USER_AGENT},
            timeout=12,
        )
        r.raise_for_status()
        js = r.json()
        return [{"lat": float(it["lat"]), "lon": float(it["lon"]),
                 "display": it.get("display_name", q)} for it in js]

    @st.cache_data(ttl=24*3600, show_spinner=False)
    def _mapsco_cached(q: str, limit: int) -> List[Dict]:
        r = requests.get(MAPSCO_URL, params={"q": q, "limit": str(limit)},
                         headers={"User-Agent": USER_AGENT}, timeout=12)
        r.raise_for_status()
        js = r.json() if isinstance(r.json(), list) else []
        out: List[Dict] = []
        for it in js:
            lat  = it.get("lat")
            lon  = it.get("lon")
            disp = it.get("display_name") or it.get("name") or q
            if lat and lon:
                out.append({"lat": float(lat), "lon": float(lon), "display": disp})
        return out
else:
    def _locationiq_cached(q, limit, key): return []
    def _nominatim_cached(q, limit): return []
    def _mapsco_cached(q, limit): return []


def geocode_candidates(q: str, limit: int = 6) -> List[Dict]:
    """
    Încearcă providerii în ordine: LocationIQ → Nominatim → maps.co.
    La primul răspuns valid, se oprește și îl returnează.
    """
    q_eff = (q or "").strip()
    if not q_eff:
        return []

    last_err: Optional[str] = None

    if LOCATIONIQ_KEY:
        try:
            out = _locationiq_cached(q_eff, int(limit), LOCATIONIQ_KEY)
            if st is not None:
                st.session_state["_geocode_source"] = "LocationIQ"
                st.session_state.pop("_geocode_error", None)
            return out
        except Exception as e:
            last_err = f"LocationIQ: {e}"

    try:
        out = _nominatim_cached(q_eff, int(limit))
        if st is not None:
            st.session_state["_geocode_source"] = "Nominatim"
            st.session_state.pop("_geocode_error", None)
        return out
    except Exception as e:
        last_err = f"Nominatim: {e}"

    try:
        out2 = _mapsco_cached(q_eff, int(limit))
        if out2:
            if st is not None:
                st.session_state["_geocode_source"] = "maps.co"
                st.session_state.pop("_geocode_error", None)
            return out2
    except Exception as e2:
        last_err = last_err or f"maps.co: {e2}"

    if st is not None and last_err:
        st.session_state["_geocode_error"] = (
            f"Geocodarea nu este disponibilă momentan. ({last_err})"
        )
    return []

# ---------------- Rutare OSRM (Multi-Punct) ----------------
if st is not None:
    @st.cache_data(ttl=24*3600, show_spinner=False)
    def _route_osrm_multi_cached(pts_coords: List[Tuple[float, float]]) -> Optional[Dict]:
        coord_str = ";".join([f"{lon},{lat}" for lat, lon in pts_coords])
        url = (
            f"https://router.project-osrm.org/route/v1/driving/{coord_str}"
            f"?overview=full&alternatives=false&steps=false&geometries=geojson"
        )
        r = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=20)

        if r.status_code == 400:
            data = r.json()
            if data.get("code") == "NoRoute":
                return {"error": "Nu s-a găsit un traseu auto între aceste puncte pe hartă."}
            return {"error": f"Eroare OSRM: {data.get('code', 'Necunoscută')}"}

        r.raise_for_status()
        data   = r.json()
        routes = data.get("routes") or []
        if not routes:
            return None

        route   = routes[0]
        legs    = route.get("legs", [])
        legs_km = [leg["distance"] / 1000.0 for leg in legs]

        geom   = route.get("geometry", {})
        coords = []
        if geom and geom.get("type") == "LineString":
            coords = geom.get("coordinates") or []

        return {"legs_km": legs_km, "coords": coords}
else:
    def _route_osrm_multi_cached(pts_coords): return None


def route_osrm_multi(pts_coords: List[Tuple[float, float]]) -> Optional[Dict]:
    """
    Apelează OSRM pentru toate punctele într-un singur request.
    Nu face sleep; nu scrie pe disc. Cache-ul e @st.cache_data.
    Facem o copie a rezultatului înainte de simplificare ca să nu modificăm
    obiectul returnat din cache.
    """
    try:
        res = _route_osrm_multi_cached(pts_coords)
        if res is not None and not res.get("error") and res.get("coords"):
            res = dict(res)
            res["coords"] = _simplify_coords(res["coords"], PATH_SUBSAMPLE_STEP)
        return res
    except Exception as e:
        return {"error": f"Eroare de rețea la calculul rutei: {e}"}


def route_osrm_multi_retry(pts_coords: List[Tuple[float, float]], tries: int = 2) -> Optional[Dict]:
    """Sleep-ul de 0.3s de aici este acceptabil: se execută doar pe eroare, nu pe request normal."""
    res = None
    for _ in range(max(1, tries)):
        res = route_osrm_multi(pts_coords)
        if res and "error" not in res:
            return res
        if res and "error" in res and "Nu s-a găsit" in res["error"]:
            return res
        time.sleep(0.3)
    return res

# ---------------- Favorite (per-sesiune + export/import JSON) ----------------
# Motivul schimbării față de v6.5:
# _FAV_LOCAL scria în ~/.foaieparcurs_fav.json pe serverul Streamlit Cloud,
# ceea ce înseamnă că (1) toți utilizatorii împărțeau aceleași favorite și
# (2) datele se pierdeau la fiecare deployment.
# Soluția: favorite în st.session_state (izolate per utilizator, per sesiune)
# + export/import JSON pentru persistență pe dispozitivul utilizatorului.

def _fav_all() -> Dict[str, Dict]:
    return st.session_state.setdefault("favorites", {})

def _fav_save(name: str, payload: Dict) -> None:
    _fav_all()[name] = payload

def _fav_delete(name: str) -> None:
    _fav_all().pop(name, None)

def _fav_sheet_available() -> bool:
    return bool(GSPREAD_SERVICE_ACCOUNT_JSON and GOOGLE_SHEET_ID)

def _fav_sheet_append(payload: Dict) -> bool:
    if not _fav_sheet_available(): return False
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        sa_info = json.loads(GSPREAD_SERVICE_ACCOUNT_JSON)
        scopes  = ["https://www.googleapis.com/auth/spreadsheets"]
        creds   = Credentials.from_service_account_info(sa_info, scopes=scopes)
        gc      = gspread.authorize(creds)
        sh      = gc.open_by_key(GOOGLE_SHEET_ID)
        try:
            ws = sh.worksheet("favorites")
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title="favorites", rows=1000, cols=4)
            ws.append_row(["date", "name", "start", "stops_json"])
        row = [
            datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
            payload.get("name", ""),
            payload.get("start", ""),
            json.dumps(payload.get("stops", []), ensure_ascii=False),
        ]
        ws.append_row(row)
        return True
    except Exception:
        return False

def _fav_payload_from_state() -> Dict:
    start_txt = st.session_state.get("txt_start") or st.session_state.get("start") or ""
    stops_txt = []
    for key in st.session_state.get("stops_keys", []):
        stops_txt.append(st.session_state.get(f"txt_{key}") or st.session_state.get(key) or "")
    return {"name": "", "start": start_txt, "stops": stops_txt}

def _fav_apply_to_state(payload: Dict) -> None:
    st.session_state["txt_start"] = payload.get("start", "")
    st.session_state["stops_keys"] = []
    for i, stop_txt in enumerate(payload.get("stops", [])):
        key = f"stop_{i}"
        st.session_state["stops_keys"].append(key)
        st.session_state[f"txt_{key}"] = stop_txt
        for suf in ("_cands", "_sel", "_lat", "_lon", "_display", "_last_fetch_ts", "_query"):
            st.session_state.pop(f"{key}{suf}", None)

# ---------------- UI helpers ----------------
def _init_addr_state(key: str, default_text: str = "") -> None:
    if st is None: return
    if f"txt_{key}" not in st.session_state:
        st.session_state[f"txt_{key}"] = default_text
    st.session_state.setdefault(f"{key}_cands", [])
    st.session_state.setdefault(f"{key}_sel", 0)
    st.session_state.setdefault(f"{key}_lat", None)
    st.session_state.setdefault(f"{key}_lon", None)
    st.session_state.setdefault(f"{key}_display", "")

def _refresh_candidates_if_due(key: str) -> None:
    if st is None: return
    q      = (st.session_state.get(f"txt_{key}") or "").strip()
    last_q = (st.session_state.get(f"{key}_query") or "").strip()
    if q and q != last_q and len(q) >= 3:
        st.session_state.pop("_geocode_error", None)
        cands = geocode_candidates(q, limit=6)
        st.session_state[f"{key}_cands"]  = cands
        st.session_state[f"{key}_query"]  = q
        st.session_state[f"{key}_sel"]    = 0

def _move_stop(old_idx: int, new_idx: int) -> None:
    keys = st.session_state.get("stops_keys", [])
    if 0 <= old_idx < len(keys) and 0 <= new_idx < len(keys):
        keys.insert(new_idx, keys.pop(old_idx))

def _render_address_row(label: str, key: str, index: int, total: int) -> None:
    if st is None: return
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("<div class='op-row-marker'></div>", unsafe_allow_html=True)
    ctitle, cactions = st.columns([0.85, 0.15])
    with ctitle:
        st.markdown(f"<p class='card-title'>Oprire #{index+1}</p>", unsafe_allow_html=True)
    with cactions:
        cup, cdown, crm = st.columns(3)
        with cup:
            if st.button("↑", key=f"up_{key}_{index}", help="Mută în sus",
                         type="secondary", disabled=(index == 0)):
                _move_stop(index, index - 1)
                st.rerun()
        with cdown:
            if st.button("↓", key=f"down_{key}_{index}", help="Mută în jos",
                         type="secondary", disabled=(index >= total - 1)):
                _move_stop(index, index + 1)
                st.rerun()
        with crm:
            if st.button("✖", key=f"rm_{key}_{index}", help="Șterge oprirea", type="secondary"):
                st.session_state.setdefault("_to_remove", []).append(key)

    cont = st.container()
    cont.text_input(label, key=f"txt_{key}")
    _refresh_candidates_if_due(key)

    cands = st.session_state.get(f"{key}_cands", [])
    src   = st.session_state.get("_geocode_source")
    if cands:
        if src: cont.caption(f"Sugestii de la: {src}")
        labels = [c["display"] for c in cands]
        idx = cont.selectbox(
            "Alege adresa",
            options=list(range(len(labels))),
            format_func=lambda i: labels[i],
            index=st.session_state.get(f"{key}_sel", 0),
            key=f"sel_{key}",
        )
        st.session_state[f"{key}_lat"]     = cands[idx]["lat"]
        st.session_state[f"{key}_lon"]     = cands[idx]["lon"]
        st.session_state[f"{key}_display"] = cands[idx]["display"]
        st.session_state[key]              = cands[idx]["display"]
    else:
        err = st.session_state.get("_geocode_error")
        if err: cont.warning(err)
        else:   cont.caption(
            "<span class='muted'>Tastează minim 3 caractere și apasă Enter pentru sugestii.</span>",
            unsafe_allow_html=True,
        )
    st.markdown("</div>", unsafe_allow_html=True)

def _collect_point_from_state(key: str) -> Optional[Dict]:
    txt  = (st.session_state.get(f"txt_{key}") or st.session_state.get(key) or "").strip()
    lat  = st.session_state.get(f"{key}_lat")
    lon  = st.session_state.get(f"{key}_lon")
    disp = st.session_state.get(f"{key}_display") or st.session_state.get(key) or txt
    if lat and lon:
        return {"lat": float(lat), "lon": float(lon), "display": disp or txt}
    if txt:
        with st.spinner(f"Geocodare: {txt}…"):
            try:
                cands = geocode_candidates(txt, limit=1)
                if cands:
                    c = cands[0]
                    st.session_state[f"{key}_lat"]     = c["lat"]
                    st.session_state[f"{key}_lon"]     = c["lon"]
                    st.session_state[f"{key}_display"] = c["display"]
                    st.session_state[key]              = c["display"]
                    return {"lat": float(c["lat"]), "lon": float(c["lon"]), "display": c["display"]}
            except Exception:
                pass
    return None

# ---------------- Hartă (pydeck) ----------------
def _fit_view(points: List[Tuple[float, float]]) -> Tuple[float, float, float]:
    if not points: return (44.43, 26.10, 9)
    lats = [p[0] for p in points]
    lons = [p[1] for p in points]
    lat_center = (min(lats) + max(lats)) / 2.0
    lon_center = (min(lons) + max(lons)) / 2.0
    span = max(0.01, max(max(lats)-min(lats), max(lons)-min(lons)))
    if   span < 0.02: zoom = 13
    elif span < 0.05: zoom = 12
    elif span < 0.1:  zoom = 11
    elif span < 0.3:  zoom = 10
    elif span < 0.7:  zoom = 9
    elif span < 1.5:  zoom = 8
    else:             zoom = 6
    return (lat_center, lon_center, zoom)

def _render_map(all_points: List[Dict], all_paths: List[List[List[float]]]) -> None:
    try:
        import pydeck as pdk
    except ImportError:
        st.info("Instalează `pydeck` pentru hartă interactivă.")
        return

    scatter_data = [{"position": [p["lon"], p["lat"]], "label": p["display"]} for p in all_points]
    path_data    = [{"path": coords} for coords in all_paths if coords and len(coords) >= 2]
    clat, clon, zoom = _fit_view([(p["lat"], p["lon"]) for p in all_points])

    scatter = pdk.Layer("ScatterplotLayer", data=scatter_data, get_position="position",
                        get_radius=80, get_fill_color=[255, 99, 71], pickable=True, auto_highlight=True)
    path    = pdk.Layer("PathLayer", data=path_data, get_path="path",
                        get_color=[0, 122, 255], get_width=4, width_min_pixels=2)
    view    = pdk.ViewState(latitude=clat, longitude=clon, zoom=zoom)
    r       = pdk.Deck(layers=[path, scatter], initial_view_state=view, tooltip={"text": "{label}"})
    st.pydeck_chart(r, use_container_width=True)

# ---------------- APP ----------------
def run_streamlit_app() -> None:
    if st is None:
        print("Streamlit nu este disponibil.")
        return

    st.title("🚗 Foaie de parcurs")
    src_badge = st.session_state.get("_geocode_source")
    if src_badge:
        st.markdown(
            f"<span class='muted'>Sursă geocodare curentă: <b>{src_badge}</b></span>",
            unsafe_allow_html=True,
        )

    # ---- Data foii ----
    st.markdown("<h4 class='section'>🗓️ Data foii</h4>", unsafe_allow_html=True)
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.session_state.setdefault("calc_date", date.today())
    st.session_state["calc_date"] = st.date_input("Alege data foii", value=st.session_state["calc_date"])
    st.markdown("</div>", unsafe_allow_html=True)

    # ---- Favorite ----
    with st.expander("⭐ Favorite", expanded=False):
        st.markdown("<div class='card'>", unsafe_allow_html=True)
        st.caption(
            "Favoritele sunt salvate în sesiunea curentă. "
            "Exportă-le ca JSON și importă-le la sesiunea următoare pentru a le păstra."
        )

        fav_col1, fav_col2 = st.columns([0.7, 0.3])
        with fav_col1:
            fav_name = st.text_input("Nume traseu favorit")
        with fav_col2:
            if st.button("💾 Salvează", use_container_width=True):
                payload = _fav_payload_from_state()
                payload["name"] = (
                    fav_name.strip() or f"traseu-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
                )
                _fav_save(payload["name"], payload)
                saved_remote = False
                if _fav_sheet_available():
                    saved_remote = _fav_sheet_append(payload)
                if saved_remote:
                    st.success(f"„{payload['name']}" salvat în sesiune și în Google Sheets.")
                else:
                    st.success(f"„{payload['name']}" salvat în sesiunea curentă.")

        local_favs = sorted(_fav_all().keys())
        lf1, lf2, lf3 = st.columns([0.6, 0.2, 0.2])
        with lf1:
            sel_fav = st.selectbox("Alege favorit", options=["(none)"] + local_favs)
        with lf2:
            if st.button("↩️ Încarcă", use_container_width=True, disabled=(sel_fav == "(none)")):
                _fav_apply_to_state(_fav_all().get(sel_fav, {}))
                st.rerun()
        with lf3:
            if st.button("🗑️ Șterge", use_container_width=True, disabled=(sel_fav == "(none)")):
                _fav_delete(sel_fav)
                st.rerun()

        # Export / Import JSON
        exp_col, imp_col = st.columns(2)
        with exp_col:
            if _fav_all():
                fav_bytes = json.dumps(_fav_all(), ensure_ascii=False, indent=2).encode("utf-8")
                st.download_button(
                    "⬇️ Exportă favorite (JSON)",
                    data=fav_bytes,
                    file_name="favorite_trasee.json",
                    mime="application/json",
                    use_container_width=True,
                )
        with imp_col:
            uploaded = st.file_uploader(
                "📂 Importă favorite (JSON)", type=["json"], key="fav_import"
            )
            if uploaded is not None:
                try:
                    imported = json.load(uploaded)
                    if isinstance(imported, dict):
                        _fav_all().update(imported)
                        st.success(f"Importate {len(imported)} favorite.")
                    else:
                        st.error("Fișierul JSON nu are formatul așteptat (obiect cheie→traseu).")
                except Exception:
                    st.error("Fișier JSON invalid sau corupt.")

        st.markdown("</div>", unsafe_allow_html=True)

    # ---- Traseu ----
    st.markdown("<h4 class='section'>🧭 Traseu</h4>", unsafe_allow_html=True)
    st.markdown("<div class='card'>", unsafe_allow_html=True)

    st.markdown("**📍 Punct de plecare**")
    _init_addr_state("start", "Piata Unirii, Bucuresti")
    cont = st.container()
    cont.text_input("Adresa de plecare", key="txt_start")
    _refresh_candidates_if_due("start")
    start_cands = st.session_state.get("start_cands", [])
    src         = st.session_state.get("_geocode_source")
    if start_cands:
        if src: cont.caption(f"Sugestii de la: {src}")
        labels = [c["display"] for c in start_cands]
        idx = cont.selectbox(
            "Alege adresa",
            options=list(range(len(labels))),
            format_func=lambda i: labels[i],
            index=st.session_state.get("start_sel", 0),
            key="sel_start",
        )
        st.session_state["start_lat"]     = start_cands[idx]["lat"]
        st.session_state["start_lon"]     = start_cands[idx]["lon"]
        st.session_state["start_display"] = start_cands[idx]["display"]
        st.session_state["start"]         = start_cands[idx]["display"]
    else:
        err = st.session_state.get("_geocode_error")
        if err: cont.warning(err)
        else:   cont.caption(
            "<span class='muted'>Tastează minim 3 caractere și apasă Enter pentru sugestii.</span>",
            unsafe_allow_html=True,
        )

    st.markdown("**🛑 Opriri**")
    if "stops_keys" not in st.session_state:
        st.session_state.stops_keys = ["stop_0"]
        _init_addr_state("stop_0", "")

    with st.expander("➕ Adaugă mai multe opriri deodată"):
        st.session_state.setdefault("bulk_add_text", "")
        if st.session_state.get("_bulk_clear", False):
            st.session_state["bulk_add_text"] = ""
            st.session_state["_bulk_clear"]   = False
        bulk_txt = st.text_area("Introdu adrese (câte una pe linie)", key="bulk_add_text")
        if st.button("Adaugă aceste opriri", key="bulk_add_btn"):
            lines = [ln.strip() for ln in (bulk_txt or "").splitlines() if ln.strip()]
            empty_keys = [
                k for k in st.session_state.get("stops_keys", [])
                if not (st.session_state.get(f"txt_{k}") or st.session_state.get(k) or "").strip()
            ]
            i = 0
            for k in empty_keys:
                if i >= len(lines): break
                st.session_state[f"txt_{k}"] = lines[i]; i += 1
                for suf in ("_cands","_sel","_lat","_lon","_display","_last_fetch_ts","_query"):
                    st.session_state.pop(f"{k}{suf}", None)
            for ln in lines[i:]:
                new_key = f"stop_{len(st.session_state.stops_keys)}"
                st.session_state.stops_keys.append(new_key)
                _init_addr_state(new_key, ln)
            st.session_state["_bulk_clear"] = True
            st.rerun()

    st.session_state.pop("_to_remove", None)
    for idx, key in enumerate(list(st.session_state.stops_keys)):
        _init_addr_state(key)
        _render_address_row("Adresă", key, idx, len(st.session_state.stops_keys))

    act1, act2 = st.columns([0.5, 0.5])
    with act1:
        if st.button("➕ Adăugare oprire", key="add_stop_btn", use_container_width=True):
            new_key = f"stop_{len(st.session_state.stops_keys)}"
            st.session_state.stops_keys.append(new_key)
            _init_addr_state(new_key, "")
            st.rerun()
    with act2:
        st.session_state.setdefault("confirm_rm_all", False)
        st.session_state["confirm_rm_all"] = st.checkbox(
            "Confirmă ștergerea tuturor", value=False, key="confirm_rm_all_cb"
        )
        if st.button(
            "🗑️ Șterge toate opririle", key="rm_all_btn",
            use_container_width=True, disabled=not st.session_state["confirm_rm_all"]
        ):
            st.session_state["_to_remove"]   = list(st.session_state.stops_keys)
            st.session_state["confirm_rm_all"] = False

    st.markdown("**⚙️ Opțiuni traseu**")
    close_loop = st.checkbox("Revenire la punctul de plecare (închidere circuit)")
    st.session_state["close_loop"] = bool(close_loop)

    remove_list = st.session_state.pop("_to_remove", [])
    if remove_list:
        for k in remove_list:
            if k in st.session_state.stops_keys:
                st.session_state.stops_keys.remove(k)
            for suf in ("_cands","_sel","_lat","_lon","_display","_last_fetch_ts","_query"):
                st.session_state.pop(f"{k}{suf}", None)
            st.session_state.pop(f"txt_{k}", None)
        st.rerun()

    if st.button("Calculează traseul", key="calc_btn", use_container_width=True):
        issues: List[str] = []
        pts:    List[Dict] = []

        start_pt = _collect_point_from_state("start")
        if start_pt:
            pts.append(start_pt)
        else:
            issues.append("Punctul de plecare nu a putut fi geocodat.")

        for key in st.session_state.get("stops_keys", []):
            p = _collect_point_from_state(key)
            if p:
                pts.append(p)
            else:
                txt = (st.session_state.get(f"txt_{key}") or "").strip()
                issues.append(f"Oprirea „{txt or key}" nu a putut fi geocodată.")

        for msg in issues:
            st.warning(msg)

        if len(pts) < 2:
            st.error("Adaugă minim o oprire validă.")
        else:
            if st.session_state.get("close_loop"):
                pts.append(pts[0])

            pts_coords = [(p["lat"], p["lon"]) for p in pts]
            res        = route_osrm_multi_retry(pts_coords)

            if not res:
                st.error("Eroare neașteptată la calcularea rutei (fără răspuns de la server).")
            elif "error" in res:
                st.error(res["error"])
            else:
                legs_km = res.get("legs_km", [])
                coords  = res.get("coords", [])

                if len(legs_km) == len(pts) - 1:
                    segments = []
                    for i in range(len(pts) - 1):
                        a, b = pts[i], pts[i + 1]
                        segments.append({
                            "from": a["display"],
                            "to":   b["display"],
                            "km_oneway": km_round(float(legs_km[i]), 1),
                        })

                    st.session_state["segments"] = segments
                    st.session_state["paths"]    = [coords] if coords else []

                    if any(seg["km_oneway"] == 0 for seg in segments):
                        st.info("Atenție: unele segmente au 0 km. Punctele ar putea fi prea apropiate sau identice.")
                    st.success("Traseul a fost calculat. Poți bifa multiplicatorii mai jos.")
                else:
                    st.error("Răspuns invalid de la serviciul de hărți (numărul de segmente nu corespunde).")

    st.markdown("</div>", unsafe_allow_html=True)

    # ---- Segmente ----
    if st.session_state.get("segments"):
        st.markdown("<h4 class='section'>🧭 Segmente</h4>", unsafe_allow_html=True)
        segments    = st.session_state["segments"]
        data_foaie  = st.session_state.get("calc_date", date.today())
        total       = 0.0
        rows        = []
        points_for_map: List[Dict] = []

        if st.session_state.get("start_lat") and st.session_state.get("start_lon"):
            points_for_map.append({
                "lat":     float(st.session_state["start_lat"]),
                "lon":     float(st.session_state["start_lon"]),
                "display": st.session_state.get("start") or "Start",
            })
        for key in st.session_state.get("stops_keys", []):
            lat  = st.session_state.get(f"{key}_lat")
            lon  = st.session_state.get(f"{key}_lon")
            disp = st.session_state.get(f"{key}_display") or st.session_state.get(key)
            if lat and lon:
                points_for_map.append({"lat": float(lat), "lon": float(lon), "display": disp or "Oprire"})

        for i, seg in enumerate(segments):
            c1, c2 = st.columns([0.7, 0.3])
            with c1:
                st.markdown(f"• <b>{seg['from']}</b> → <b>{seg['to']}</b>", unsafe_allow_html=True)
            with c2:
                checked = st.checkbox("dus-întors", key=f"seg_rt_{i}",
                                      value=st.session_state.get(f"seg_rt_{i}", False))
                reps    = st.number_input("×", min_value=1, max_value=50, step=1,
                                          key=f"seg_rep_{i}", value=st.session_state.get(f"seg_rep_{i}", 1))
            effective = seg["km_oneway"] * (2 if checked else 1) * int(reps)
            total    += effective
            st.markdown(
                f"<span class='muted'>Distanță (×{int(reps)}): <b>{effective} km</b></span>",
                unsafe_allow_html=True,
            )
            rows.append({
                "Data":           data_foaie.strftime("%d.%m.%Y"),
                "Plecare":        seg["from"],
                "Destinație":     seg["to"],
                "Dus-întors":     "Da" if checked else "Nu",
                "Înmulțiri (×)":  int(reps),
                "Km parcurși":    effective,
            })

        st.success(f"Total km: {total}")

        st.markdown("<h4 class='section'>🗺️ Hartă traseu</h4>", unsafe_allow_html=True)
        _render_map(points_for_map, st.session_state.get("paths", []))

        df = pd.DataFrame(rows)
        st.dataframe(df, use_container_width=True)

        csv_bytes = df.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "⬇️ Descarcă CSV", csv_bytes,
            file_name=f"foaie_parcurs_{data_foaie.strftime('%Y%m%d')}.csv",
            mime="text/csv", use_container_width=True,
        )

        bio = io.BytesIO()
        try:
            from openpyxl.styles import Font
            with pd.ExcelWriter(bio, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, header=False,
                            sheet_name="Foaie de parcurs", startrow=1)
                ws = writer.sheets["Foaie de parcurs"]
                for col_idx, col_name in enumerate(df.columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name).font = Font(bold=True)
                tot_col = df.shape[1] + 1
                ws.cell(row=1, column=tot_col, value="KM totali").font = Font(bold=True)
                ws.cell(row=2, column=tot_col, value=float(total)).font = Font(bold=True)
                try:
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(tot_col)
                    ws.column_dimensions[col_letter].width = 15
                    ws.cell(row=2, column=tot_col).number_format = "0.0"
                except Exception:
                    pass
                ws.freeze_panes = "A2"
            bio.seek(0)
            st.download_button(
                "⬇️ Descarcă Excel", bio.getvalue(),
                file_name=f"foaie_parcurs_{data_foaie.strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as ex:
            st.warning("Nu am putut genera Excel. Verifică `openpyxl`. Detalii:")
            st.exception(ex)
            st.info("CSV rămâne disponibil.")

# ---------------- Teste ----------------
def _run_basic_tests() -> None:
    # km_round
    assert km_round(12.34, 1) == 12.3, "km_round 12.34 → 12.3"
    assert km_round(12.35, 1) in (12.3, 12.4), "km_round 12.35 edge"
    assert km_round(0.0, 1) == 0.0, "km_round 0"
    assert km_round(100.0, 1) == 100.0, "km_round 100"

    # _simplify_coords
    coords = [[i, i] for i in range(10)]
    simplified = _simplify_coords(coords, step=3)
    assert simplified[-1] == coords[-1], "simplify_coords păstrează ultimul punct"
    assert len(simplified) < len(coords), "simplify_coords reduce lungimea"

    # DataFrame structure
    df = pd.DataFrame([{
        "Data": "01.01.2025", "Plecare": "A", "Destinație": "B",
        "Dus-întors": "Nu", "Înmulțiri (×)": 2, "Km parcurși": 24.6,
    }])
    assert list(df.columns) == ["Data", "Plecare", "Destinație", "Dus-întors", "Înmulțiri (×)", "Km parcurși"]

    # _round5
    assert _round5(1.123456789) == 1.12346, "_round5"

    print("Toate testele au trecut.")

if __name__ == "__main__":
    if "--test" in sys.argv:
        _run_basic_tests()
        sys.exit(0)
    if st is not None:
        run_streamlit_app()
    else:
        print("Folosește: streamlit run app.py")
